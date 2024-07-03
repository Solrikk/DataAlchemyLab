import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as openpyxlImage
from PIL import Image
import io


def save_images_from_excel(excel_path, output_folder):
  if not os.path.exists(output_folder):
    os.makedirs(output_folder)

  workbook = load_workbook(excel_path, data_only=True)

  for sheetname in workbook.sheetnames:
    worksheet = workbook[sheetname]

    for img_num, image in enumerate(worksheet._images):
      img_bytes = image._data()

      img = Image.open(io.BytesIO(img_bytes))

      cell_location = image.anchor._from
      cell_row = cell_location.row + 1
      cell_column = cell_location.col + 1
      cell_address = f'{cell_column}{cell_row}'

      img_filename = os.path.join(
          output_folder, f'{sheetname}_{cell_address}_image{img_num + 1}.png')

      img.save(img_filename)


excel_path = 'Foto.xlsx'
output_folder = 'output_images'

save_images_from_excel(excel_path, output_folder)
